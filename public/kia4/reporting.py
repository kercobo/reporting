#!/usr/bin/python -tt
import mysql.connector
import sys, getopt
import yaml
import openpyxl

USERNAME = 'root'
PASSWORD = 'bismillah'
HOST = 'localhost'
DATABASE = 'analytics'


def query_mysql(query_string):
    print "Try to open sql..."
    cnx = mysql.connector.connect(user=USERNAME, password=PASSWORD, database=DATABASE, host=HOST, port=3306,
                                  buffered=True, connection_timeout=500)
    cursor = cnx.cursor()
    query = query_string
    cursor.execute(query)

    print "Successfully opened sql"

    result = dict()

    for control, value in cursor:
        result[str(control).replace("_", " ")] = value
        result[str(control).replace("+", " ")] = value

    cursor.close()
    cnx.close()
    return result


def read_json(filename):
    try:
        print "Try to open " + filename + " file..."
        with open(filename) as data_file:
            data = yaml.safe_load(data_file)
        print "Successfully opened " + filename + " file"
    except IOError as e:
        print e
        sys.exit()

    return data


def map_to_xls(query, xls, qval, qcontrol, xlsfilename):

    wb = openpyxl.load_workbook(filename=xlsfilename)
    ws = wb.active

    if qcontrol == xls["control"]:
        for num in range(int(xls['rowStart']), int(xls['rowEnd'])):
            control_name = str(ws[xls['col']+str(num)].value)
            if control_name in query and qval in xls['parameter']:
                ws[xls['parameter'][qval] + str(num)] = query[control_name]

    wb.save(filename=xlsfilename)

    return


def main():
    # Handle argument
    query_list_file = ''
    xls_map_file = ''
    xls_tempate = ''

    try:
        opts, args = getopt.getopt(sys.argv[1:], "hc:j:x:", ["ifile=","ofile=","xfile="])
    except getopt.GetoptError:
        print 'reporting.py -c <queryjson> -j <xlsjson> -x <xlstemplate>'
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print 'reporting.py -c <queryjson> -j <xlsjson> -x <xlstemplate>'
            sys.exit()
        elif opt in ("-c", "--ifile"):
            query_list_file = arg
        elif opt in ("-j", "--ofile"):
            xls_map_file = arg
        elif opt in ("-x", "--xfile"):
            xls_tempate = arg

    if not query_list_file or not xls_map_file or not xls_tempate:
        print 'reporting.py -c <queryjson> -j <xlsjson> -x <xlstemplate>'
        sys.exit()

    query_list = read_json(query_list_file)
    xls_def = read_json(xls_map_file)

    result_query = []

    for idx, val in enumerate(query_list):
        r = dict()
        r["query"] = (query_mysql(val['query']))
        r["val"] = val['value']
        r["control"] = val['control']
        result_query.append(r)

    for xls in xls_def:
        for query in result_query:
            map_to_xls(query['query'], xls, query['val'], query['control'], xls_tempate)


if __name__ == '__main__':
    main()
